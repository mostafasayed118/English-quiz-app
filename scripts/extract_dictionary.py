"""
extract_dictionary.py
---------------------
Extracts all meaningful English words from the question bank, translates them
to Arabic via Google Translate (batch mode), and writes an Excel file.

    | #  | English       | Arabic        | Categories          | Frequency |
    | 1  | authenticate  | مصادقة        | IT & Technology     | 3         |

Run from the project root:
    python scripts/extract_dictionary.py

Output: data/english_arabic_dictionary.xlsx
"""

import json
import re
import sys
import time
from collections import Counter
from pathlib import Path

from deep_translator import GoogleTranslator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
QUESTIONS_FILE = PROJECT_ROOT / "data" / "questions.json"
OUTPUT_FILE = PROJECT_ROOT / "data" / "english_arabic_dictionary.xlsx"

MIN_WORD_LEN = 3

# Common English stop words
STOP_WORDS = {
    "the","and","for","are","but","not","you","all","can","had","her","was",
    "one","our","out","has","his","how","its","may","new","now","old","see",
    "way","who","did","get","let","say","she","too","use","any","him","been",
    "from","have","this","that","with","they","will","each","make","like",
    "than","them","then","what","when","your","also","more","very","does",
    "some","time","such","into","just","only","other","most","over","after",
    "before","should","would","could","might","being","these","those",
    "where","while","about","there","their","which","through","between",
    "because","during","without","again","further","both","few","own","same",
    "here","why","yet","already","still","often","well","even","ever","never",
    "always","under","above","below","against","among","around","behind",
    "beside","besides","beyond","inside","outside","within","despite","since",
    "until","upon","rather","though","although","hence","thus","therefore",
    "however","nevertheless","meanwhile","otherwise","instead","indeed",
    "must","shall","need","dare","ought","used","able","going","take","come",
    "know","think","first","back","give","much","many","part","want","long",
    "look","help","line","turn","move","live","real","left","great","next",
    "life","hand","high","keep","last","begin","small","start","read","put",
    "set","try","ask","feel","three","state","work","year","thing","every",
    "point","find","tell","seem","mean","call","end","side","case","form",
    "day","name","happen","show","number","people","water","place","world",
    "change","play","still","learn","head","stand","letter","meet","close",
    "power","land","draw","plan","city","group","carry","ride","type","kind",
    "sort","lot","bit","half","whole","twice","once","run","right","left",
    "along","across","during","since","until","per","via","etc","e.g","i.e",
    "vs","etc","ok","yes","no","if","or","so","do","it","an","as","at","be",
    "by","up","to","in","on","of","is","no","go","us","am","um","ah","oh",
}

WORD_RE = re.compile(r"[a-zA-Z]{" + str(MIN_WORD_LEN) + r",}")

translator = GoogleTranslator(source="en", target="ar")


def extract_words(questions: list[dict]) -> tuple[list[str], dict[str, Counter]]:
    """Extract unique English words → (sorted_list, {word: Counter({cat: n})})."""
    word_cats: dict[str, Counter] = {}
    for q in questions:
        text = " ".join([
            q.get("question_text", ""),
            " ".join(q.get("options", [])),
            q.get("explanation", ""),
            q.get("category", ""),
            " ".join(q.get("tags", [])),
        ])
        text_clean = re.sub(r"[A-Z]\)\s*", " ", text)
        words = WORD_RE.findall(text_clean)
        cat = q.get("category", "Unknown")
        for w in words:
            lo = w.lower()
            if lo in STOP_WORDS:
                continue
            word_cats.setdefault(lo, Counter())[cat] += 1
    sorted_words = sorted(word_cats.keys(), key=lambda w: (-sum(word_cats[w].values()), w))
    return sorted_words, word_cats


def translate_batch(words: list[str], batch_size: int = 40) -> dict[str, str]:
    """Translate words in batches. Returns {english: arabic}."""
    result: dict[str, str] = {}
    total = len(words)
    for i in range(0, total, batch_size):
        batch = words[i : i + batch_size]
        print(f"  Batch {i // batch_size + 1}/{(total + batch_size - 1) // batch_size}: "
              f"words {i + 1}–{i + len(batch)} of {total}...")
        try:
            # translate_batch returns a list of translated strings
            translated = translator.translate_batch(batch)
            for en, ar in zip(batch, translated):
                result[en] = ar if ar else en
        except Exception as e:
            print(f"  ⚠ Batch failed ({e}), falling back to word-by-word...")
            for w in batch:
                try:
                    result[w] = translator.translate(w) or w
                except:
                    result[w] = w
                time.sleep(0.15)
        # Pause between batches to stay under rate limits
        time.sleep(1.0)
    return result


def write_excel(words, translations, categories, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "English-Arabic Dictionary"

    hdr_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    en_font  = Font(name="Calibri", size=11, color="1E293B")
    ar_font  = Font(name="Arial", size=11, color="1E293B")
    en_align = Alignment(horizontal="left", vertical="center")
    ar_align = Alignment(horizontal="right", vertical="center", reading_order=2)
    cat_font = Font(name="Calibri", size=10, color="64748B")
    border   = Border(
        left=Side("thin","CBD5E1"), right=Side("thin","CBD5E1"),
        top=Side("thin","CBD5E1"),  bottom=Side("thin","CBD5E1"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col, h in enumerate(["#", "English", "Arabic", "Categories", "Frequency"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 28

    for i, word in enumerate(words, 1):
        ar = translations.get(word, word)
        cats = ", ".join(categories[word].keys())
        freq = sum(categories[word].values())
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=word).font = en_font
        ws.cell(row=row, column=2).alignment = en_align
        ws.cell(row=row, column=3, value=ar).font = ar_font
        ws.cell(row=row, column=3).alignment = ar_align
        ws.cell(row=row, column=4, value=cats).font = cat_font
        ws.cell(row=row, column=4).alignment = en_align
        ws.cell(row=row, column=5, value=freq).alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = alt_fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    print(f"\n  ✅ Saved: {output}  ({len(words)} words, {output.stat().st_size / 1024:.1f} KB)")


def main():
    print("=" * 60)
    print("  English-Arabic Dictionary Extractor")
    print("=" * 60)

    print(f"\n[1] Load questions")
    with open(QUESTIONS_FILE, encoding="utf-8") as f:
        questions = json.load(f)
    print(f"  {len(questions)} questions loaded")

    print(f"\n[2] Extract unique words")
    words, cats = extract_words(questions)
    print(f"  {len(words)} unique words (after stop-word removal)")

    print(f"\n[3] Translate to Arabic (batch mode)")
    trans = translate_batch(words)

    print(f"\n[4] Write Excel")
    write_excel(words, trans, cats, OUTPUT_FILE)

    print(f"\n{'=' * 60}")
    print(f"  Done!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
