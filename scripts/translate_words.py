"""
translate_words.py
------------------
Step 2: Reads data/words_raw.csv, translates each English word to Arabic
        using the `translators` library (free, no API key), and writes a
        polished Excel dictionary file.

Run from project root:
    python scripts/translate_words.py

Output: data/english_arabic_dictionary.xlsx
"""

import csv
import time
import sys
from pathlib import Path

import translators as ts
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_CSV    = PROJECT_ROOT / "data" / "words_raw.csv"
OUTPUT_XLSX  = PROJECT_ROOT / "data" / "english_arabic_dictionary.xlsx"


def translate_word(word: str) -> str:
    """Translate a single English word to Arabic using translators."""
    try:
        result = ts.translate_text(word, from_language="en", to_language="ar")
        return result if result else word
    except Exception:
        return word


def translate_batch_with_retry(words: list[str], max_retries: int = 3) -> dict[str, str]:
    """Translate words with automatic retry on failure."""
    result: dict[str, str] = {}
    total = len(words)

    for i, word in enumerate(words, 1):
        # Progress every 50 words
        if i % 50 == 0 or i == total:
            print(f"  [{i}/{total}] ({100*i//total}%) translating...", flush=True)

        for attempt in range(max_retries):
            try:
                translated = ts.translate_text(word, from_language="en", to_language="ar")
                result[word] = translated if translated else word
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    # Exponential backoff: 1s, 2s, 4s
                    time.sleep(2 ** attempt)
                else:
                    # All retries failed — keep English as placeholder
                    result[word] = word

        # Small delay between words to avoid rate-limiting
        time.sleep(0.05)

    return result


def write_excel(words, translations, categories, frequencies, output):
    """Write a formatted Excel file with the dictionary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "English-Arabic Dictionary"

    # Styles
    hdr_font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    en_font   = Font(name="Calibri", size=11, color="1E293B")
    ar_font   = Font(name="Arial", size=11, color="1E293B")
    en_align  = Alignment(horizontal="left", vertical="center")
    ar_align  = Alignment(horizontal="right", vertical="center", reading_order=2)
    cat_font  = Font(name="Calibri", size=10, color="64748B")
    freq_font = Font(name="Calibri", size=11, color="4F46E5", bold=True)
    border    = Border(
        left=Side("thin","CBD5E1"), right=Side("thin","CBD5E1"),
        top=Side("thin","CBD5E1"),  bottom=Side("thin","CBD5E1"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Headers
    for col, h in enumerate(["#", "English", "Arabic", "Categories", "Freq"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 28

    # Data
    for i, word in enumerate(words, 1):
        ar  = translations.get(word, word)
        cat = categories.get(word, "")
        freq = frequencies.get(word, 0)
        row = i + 1

        ws.cell(row=row, column=1, value=i).font = en_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=word).font = en_font
        ws.cell(row=row, column=2).alignment = en_align
        ws.cell(row=row, column=3, value=ar).font = ar_font
        ws.cell(row=row, column=3).alignment = ar_align
        ws.cell(row=row, column=4, value=cat).font = cat_font
        ws.cell(row=row, column=4).alignment = en_align
        ws.cell(row=row, column=5, value=freq).font = freq_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = alt_fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 8

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output)
    size_kb = output.stat().st_size / 1024
    print(f"\n  ✅ Saved: {output}  ({len(words)} words, {size_kb:.1f} KB)")


def main():
    print("=" * 60)
    print("  Translate English words → Arabic")
    print("=" * 60)

    # 1. Read CSV
    print(f"\n[1] Read {INPUT_CSV}")
    words = []
    categories = {}
    frequencies = {}
    with open(INPUT_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            w = row["English"]
            words.append(w)
            categories[w] = row.get("Categories", "")
            frequencies[w] = int(row.get("Frequency", 0))
    print(f"  {len(words)} words to translate")

    # 2. Translate (with progress)
    print(f"\n[2] Translate to Arabic (this may take 5–15 minutes for 3000+ words)")
    translations = translate_batch_with_retry(words)

    # 3. Write Excel
    print(f"\n[3] Write Excel")
    write_excel(words, translations, categories, frequencies, OUTPUT_XLSX)

    print(f"\n{'='*60}")
    print(f"  Done! Open {OUTPUT_XLSX.name} in Excel.")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
