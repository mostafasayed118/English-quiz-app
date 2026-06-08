"""
translate_top_words.py
----------------------
Translate only the top 200 most frequent words from the question bank.
This runs in ~1 minute. The full word list is also written for manual
completion in Excel.

Output: data/english_arabic_dictionary.xlsx (with top-200 translated)
"""

import csv
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import translators as ts

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_CSV    = PROJECT_ROOT / "data" / "words_raw.csv"
OUTPUT_XLSX  = PROJECT_ROOT / "data" / "english_arabic_dictionary.xlsx"
TOP_N = 500  # translate only the top N most frequent words


def translate_with_retry(word: str, retries: int = 3) -> str:
    for attempt in range(retries):
        try:
            r = ts.translate_text(word, from_language="en", to_language="ar")
            return r if r else word
        except Exception:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return word


def write_excel(words, translations, categories, frequencies, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "English-Arabic Dictionary"

    hdr_font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    en_font   = Font(name="Calibri", size=11, color="1E293B")
    ar_font   = Font(name="Arial", size=11, color="1E293B")
    en_align  = Alignment(horizontal="left", vertical="center")
    ar_align  = Alignment(horizontal="right", vertical="center", readingOrder=2)
    cat_font  = Font(name="Calibri", size=10, color="64748B")
    freq_font = Font(name="Calibri", size=11, color="4F46E5", bold=True)
    note_font = Font(name="Calibri", size=10, color="94A3B8", italic=True)
    border    = Border(
        left=Side("thin","CBD5E1"), right=Side("thin","CBD5E1"),
        top=Side("thin","CBD5E1"),  bottom=Side("thin","CBD5E1"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col, h in enumerate(["#", "English", "Arabic", "Categories", "Freq"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 28

    for i, word in enumerate(words, 1):
        ar   = translations.get(word, "")
        cat  = categories.get(word, "")
        freq = frequencies.get(word, 0)
        row  = i + 1

        ws.cell(row=row, column=1, value=i).font = en_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=word).font = en_font
        ws.cell(row=row, column=2).alignment = en_align
        if ar:
            ws.cell(row=row, column=3, value=ar).font = ar_font
        else:
            ws.cell(row=row, column=3, value="← add Arabic here").font = note_font
        ws.cell(row=row, column=3).alignment = ar_align
        ws.cell(row=row, column=4, value=cat).font = cat_font
        ws.cell(row=row, column=4).alignment = en_align
        ws.cell(row=row, column=5, value=freq).font = freq_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = alt_fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 8
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    print(f"  ✅ Saved: {output}  ({len(words)} words, {output.stat().st_size / 1024:.1f} KB)")


def main():
    print("=" * 60)
    print("  Translate Top 200 Words → Arabic")
    print("=" * 60)

    # Read CSV
    print(f"\n[1] Read words")
    all_words = []
    cats, freqs = {}, {}
    with open(INPUT_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            w = row["English"]
            all_words.append(w)
            cats[w] = row.get("Categories", "")
            freqs[w] = int(row.get("Frequency", 0))
    print(f"  Total words: {len(all_words)}")

    # Top N
    top = all_words[:TOP_N]
    print(f"\n[2] Translate top {TOP_N} words (most frequent)")

    translated: dict[str, str] = {}
    for i, word in enumerate(top, 1):
        if i % 25 == 0 or i == TOP_N:
            print(f"  [{i}/{TOP_N}] ({100*i//TOP_N}%)", flush=True)
        translated[word] = translate_with_retry(word)
        time.sleep(0.05)

    print(f"  Translated {len(translated)} words")

    # Write Excel: top 200 translated + remaining untranslated
    print(f"\n[3] Write Excel")
    write_excel(all_words, translated, cats, freqs, OUTPUT_XLSX)

    print(f"\n{'='*60}")
    print(f"  Done! Top {TOP_N} words are translated.")
    print(f"  The remaining words show '← add Arabic here' for you to fill.")
    print(f"  Open {OUTPUT_XLSX.name} in Excel to complete the dictionary.")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
